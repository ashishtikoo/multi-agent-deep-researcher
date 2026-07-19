"""
File Reader Utility – Reads various file formats for document-based research.
"""

import csv
import json
from pathlib import Path
from typing import Optional


def read_text_file(file_path: str) -> str:
    """Read a plain text file."""
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


def read_csv_file(file_path: str) -> str:
    """Read a CSV file and return formatted text."""
    rows = []
    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader, None)
        if headers:
            rows.append(" | ".join(headers))
            rows.append("-" * len(" | ".join(headers)))
        for row in reader:
            rows.append(" | ".join(row))
    return "\n".join(rows)


def read_json_file(file_path: str) -> str:
    """Read a JSON file and return formatted text."""
    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return json.dumps(data, indent=2, ensure_ascii=False)


def read_pdf_file(file_path: str) -> str:
    """Read a PDF file and extract text content."""
    try:
        import pypdf
    except ImportError:
        return "ERROR: pypdf not installed. Run: pip install pypdf"

    text_parts = []
    with open(file_path, "rb") as f:
        reader = pypdf.PdfReader(f)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    return "\n\n".join(text_parts)


def read_docx_file(file_path: str) -> str:
    """Read a DOCX file and extract text content."""
    try:
        import docx
    except ImportError:
        return "ERROR: python-docx not installed. Run: pip install python-docx"

    doc = docx.Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def read_xlsx_file(file_path: str) -> str:
    """Read an XLSX file and return formatted text."""
    try:
        import openpyxl
    except ImportError:
        return "ERROR: openpyxl not installed. Run: pip install openpyxl"

    text_parts = []
    wb = openpyxl.load_workbook(file_path, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            text_parts.append(" | ".join(c for c in cells if c))
        text_parts.append("")
    wb.close()
    return "\n".join(text_parts)


def read_image_file(file_path: str) -> str:
    """Read an image file and return metadata (text content extraction requires OCR)."""
    return (
        f"[Image file: {Path(file_path).name}]\n"
        "Note: Image content requires OCR for text extraction. "
        "The file has been recognized and included as a reference."
    )


def read_file(file_path: str) -> str:
    """
    Read a file based on its extension and return text content.

    Supported formats:
    - .txt, .text, .md, .markdown, .rst -> plain text
    - .csv -> CSV formatted as text
    - .json -> JSON formatted as text
    - .pdf -> PDF text extraction (requires pypdf)
    - .docx, .doc -> Word document text extraction (requires python-docx)
    - .xlsx, .xls -> Excel spreadsheet (requires openpyxl)
    - .png, .jpg, .jpeg, .gif, .bmp, .tiff -> Image (metadata only)
    """
    ext = Path(file_path).suffix.lower()

    readers = {
        ".txt": read_text_file,
        ".text": read_text_file,
        ".md": read_text_file,
        ".markdown": read_text_file,
        ".rst": read_text_file,
        ".csv": read_csv_file,
        ".json": read_json_file,
        ".pdf": read_pdf_file,
        ".docx": read_docx_file,
        ".doc": read_docx_file,
        ".xlsx": read_xlsx_file,
        ".xls": read_xlsx_file,
        ".png": read_image_file,
        ".jpg": read_image_file,
        ".jpeg": read_image_file,
        ".gif": read_image_file,
        ".bmp": read_image_file,
        ".tiff": read_image_file,
        ".tif": read_image_file,
    }

    reader = readers.get(ext)
    if reader is None:
        # Try reading as plain text as fallback
        try:
            return read_text_file(file_path)
        except Exception:
            return f"[Unsupported file type: {ext}]"

    try:
        content = reader(file_path)
        return content if content else f"[Empty {ext} file]"
    except Exception as e:
        return f"[Error reading {ext} file: {str(e)}]"
